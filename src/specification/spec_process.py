import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, NamedStyle, Font
from openpyxl.styles.borders import Border, Side
import copy

class Spec:
    def __init__(self, template_file_path, spec_gen, spec_gen_dir, additional_domains, append_message):
        self.file_path = template_file_path
        self.file_dir = spec_gen_dir
        self.spec_gen = spec_gen
        self.additional_domains = [domain.strip().upper() for domain in additional_domains if len(domain.strip()) > 0]
        self.domains = []
        self.message = append_message
        self.df_srdm = None
        self.nextgen = None
        self.sdtm32 = None
        self.sdtm33 = None
        self.sdtm32_02 = None
        self.output_path = os.path.join(self.file_dir, "VOL3SRDM_SDTM_MAPPING_SPECIFICATION_"+self.spec_gen+".xlsx")
#         self.book = load_workbook(file_path)
#         self.xlsx = pd.ExcelFile(file_path, engine="openpyxl")

    def get_domains(self):
        srdm_list = [filename for filename in os.listdir(self.file_dir) if filename.startswith(self.spec_gen) and filename.endswith('.xlsx')]
        print(srdm_list)
        df_srdm=[]
        cols = [17, 14, 19, 13]

        for i in range(len(srdm_list)):
            srdm00 = srdm_list[i]
            print(os.path.join(self.file_dir, srdm_list[i]))
            srdm01 = pd.read_excel(os.path.join(self.file_dir, srdm_list[i]), engine="openpyxl", sheet_name=3, skiprows=1, usecols=cols, names=['SType','SLength','SName','SLabel'])
            srdm01 = srdm01.loc[srdm01['SName'].str.contains('_')]
            srdm01['SType']= np.where(srdm01['SType'].str.upper() == 'VARCHAR2','Character',srdm01['SType'].str.title())
            n_col = (srdm01['SName'].str.split("_").apply(len)).max()
            collis = []
            for i in range(n_col):
                collis.append('V'+str(i))
            srdm01[collis] = srdm01['SName'].str.split('_', expand=True)
            df_srdm.append(srdm01)
        self.df_srdm = pd.concat(df_srdm)
        self.domainlist = np.unique([x[:2] for x in self.df_srdm['SName'] if isinstance(x, str) and x.count('_') == 1]).tolist()
        self.domains = np.unique([x[:2] for x in self.df_srdm['SName'] if isinstance(x, str) and x.count('_') == 1]).tolist()
        self.domains = self.domains + self.additional_domains

        self.df_srdm1 = self.df_srdm[((self.df_srdm['SName'].str.count('_')==1) | (self.df_srdm['V2'].str.contains('DTS'))) & (self.df_srdm['SType']!='Date')]

        unique_chars = lambda x: ' '.join(x.unique())
        unique_max = lambda x: x.max()
        self.df_srdm2 = self.df_srdm1.groupby('V0').agg({'SName': unique_chars,'SType': unique_chars,'SLabel': unique_chars, 'SLength':unique_max}).reset_index()
        self.df_srdm2['VCount'] = self.df_srdm2['SName'].str.count(' ') + 1
        self.message("Domains Identified are: " + str(self.domains))

    def read_nextgen(self):
        self.message("Reading NextGEN SDTM Metadata.xlsx")
        file_path = os.path.join(self.file_dir, "NextGEN SDTM Metadata.xlsx")
        self.nextgen = pd.read_excel(file_path, sheet_name="Elements", usecols=['Order','Dataset','Name','Core'])
        self.nextgen = self.nextgen[self.nextgen['Dataset'].isin(self.domains)]
        self.nextgen.reset_index(drop=True)
        self.nextgen.head()
        self.message("Success Reading NextGEN SDTM Metadata.xlsx")

    def read_sdtm_ig(self):
        self.message("Reading SDTMIG.xlsx")
        file_path = os.path.join(self.file_dir, "SDTMIG.xlsx")
        col_names = ['Version','Order','Class','Dataset','Vname','Name','Label','Type','CodeListRef','Role','Description','Core']
        self.sdtm32 = pd.read_excel(file_path, engine="openpyxl", sheet_name="SDTMIG v3.2",skiprows=1, names=col_names, usecols="A:L")
        self.sdtm33 = pd.read_excel(file_path, sheet_name="SDTMIG v3.3",skiprows=1, names=col_names, usecols="A:L")
        self.message("Success Reading SDTMIG.xlsx")

    def read_sdtm_meta(self):
        self.message("Reading Consolidated Meta.xlsx")
        file_path = os.path.join(self.file_dir, "Consolidated Meta.xlsx")
        self.sdtm32 = pd.read_excel(file_path, sheet_name="P21SDTMIG3.2_Variables"
                               ,usecols=['Order','Dataset','Name','Label', 'Type', 'Core', 'Codelist', 'Origin'])
        self.sdtm32 = self.sdtm32[self.sdtm32.Core.isin(['Required', 'Expected', 'Permissible', 'Model Permissible'])]
        self.sdtm33 = pd.read_excel(file_path, sheet_name="P21SDTMIG3.3_Variables"
                               ,usecols=['Order','Dataset','Name','Label', 'Type', 'Core', 'Codelist', 'Origin'])
        self.sdtm33 = self.sdtm33[self.sdtm33.Core.isin(['Required', 'Expected', 'Permissible', 'Model Permissible'])]
        self.labelmeta = pd.read_excel(file_path, sheet_name="MetaLabelsCompare")
        self.message("Success Reading Consolidated Meta.xlsx")

    def get_unique_domains_from_sdtm(self):
        self.message("Getting unique list of domains")
        domain_32 = self.sdtm32['Dataset'].unique()
        self.domain_32 = [x[:2] for x in domain_32 if str(x) != 'nan']

        #Get Unique list of domains in SDTM IG 3.3
        self.domain_33 = self.sdtm33['Dataset'].unique()
        self.domain_33 = [x[:2] for x in self.domain_33 if str(x) != 'nan']

        #List Domains in SDTM IG 3.3 which are not available in SDTM IG 3.2
#         self.in33_not_in32 = []
        self.in33_not_in32=[x for x in self.domain_33 if x not in self.domain_32]

        #Keep Only those domains that do not exist in 3.2 and available in SDTM IG 3.3
        self.sdtm33 = self.sdtm33[self.sdtm33.Dataset.isin(self.in33_not_in32)]

        #Consolidated SDTM IG 3.2 and 3.3 (Only those domains that do not exist in 3.2)
        self.sdtm_meta = self.sdtm32.append(self.sdtm33,ignore_index=True)

        self.fval_in32 = [x for x in self.domains if x in self.domain_32] #Domain in 3.2
        self.fval_in33 = [x for x in self.domains if x in self.in33_not_in32] #Domain in 3.3
        self.neither_32_nor_33 = [x for x in self.domains if x not in self.domain_32 + self.domain_33] #Domains in neither 3.2 nor 3.3

        self.sdtm32_02 = self.sdtm_meta[self.sdtm_meta['Dataset'].isin(self.domains)]
        self.sdtm32_02['CodeListRef']= np.where(self.sdtm32_02['Codelist'].str.upper().isin(['MEDDRA','ISO 8601','ISO 3166-1 Alpha-3']),np.nan,self.sdtm32_02['Codelist'])
        self.sdtm32_02['Type']= np.where(self.sdtm32_02['Type'].str.upper().isin(['FLOAT','INTEGER']),'Number','Character')
        self.sdtm32_02.drop('Codelist', axis=1)
        self.sdtm32_02['Core1'] = np.where(self.sdtm32_02['Core'].str.upper().isin(['EXPECTED','REQUIRED']),self.sdtm32_02['Core'].str[:3],'Perm')


        self.message("Domains in 3.2: "+str(self.fval_in32))
        self.message("Domains in 3.3: "+str(self.fval_in33))
        self.message("Domains neither in 3.2 nor in 3.3 : "+str(self.neither_32_nor_33))

    def append_data_for_all_domains(self):
        self.message("Appending data for all domains")
        self.sdtm_00=[]
        for i in range(len(self.domains)):
            self.sdtm_01 = self.sdtm_meta[self.sdtm_meta['Dataset'] == self.domains[i]]   #To filter on domain
            self.sdtm_01['Type']=np.where(self.sdtm_01['Type']=='text','Character','Number')
            self.sdtm_00.append(self.sdtm_01)
        self.sdtm_00 = pd.concat(self.sdtm_00)
        self.message("Successfully appended data")

    def set_desired_columns(self):
        self.message("Filtering desired columns")
        self.df_00 = ['All Classes']    #Create a list for filteration criteria based on 'Class' column.
        self.sdtm32_02=[]    #Create an empty list to append the data for all domains in fval_list
        for i in range(len(self.domains)):
            domain = self.domains[i]
            self.sdtm32_00 = self.sdtm_meta['Class'][self.sdtm_meta['Dataset'].isin([domain])].unique()    #To get the 'Class' of respective domain from fval.
            self.df_00.append([x+'-General' for x in self.sdtm32_00][0])    #To get the Class with corresponding '-General' for filteration
            # To filter the dataframe with required data and class in 'All Classes' and corresponding '-General' Class.
            self.sdtm32_01 = self.sdtm_meta[np.logical_and(self.sdtm_meta['Dataset'].isna(), self.sdtm_meta['Class'].isin(self.df_00))]
            self.sdtm32_01 =pd.concat([self.sdtm32_01,self.sdtm_meta[self.sdtm_meta['Dataset'] == domain]])   #To filter on domain
            self.sdtm32_01['Name']=self.sdtm32_01['Name'].str.replace('--',domain)    #replace'--' with Domain name in "Name" Column
            self.sdtm32_01['Dataset']=np.where(self.sdtm32_01['Dataset'].isnull(),domain,self.sdtm32_01['Dataset'])
            self.sdtm32_01['Core']=np.where(self.sdtm32_01['Core'].isnull(),'Perm',self.sdtm32_01['Core'])
            self.sdtm32_01['Type']=np.where(self.sdtm32_01['Type']=='Num','Number','Character')
            self.sdtm32_02.append(self.sdtm32_01)
            self.df_00=['All Classes']
        self.sdtm32_02 = pd.concat(self.sdtm32_02)
        #keep the last duplicate value per Dataset & Name
        self.sdtm32_02.drop_duplicates(subset=['Dataset','Name'], keep='last', inplace=True)
        #To Remove '()' from 'CodeListRef' Column.
        self.sdtm32_02['CodeListRef'] = self.sdtm32_02['CodeListRef'].str.replace(r'[^\w\s]+', '')
        self.sdtm32_02['CodeListRef']= np.where(self.sdtm32_02['Name'].str.upper() == 'EPOCH','EPOCH',self.sdtm32_02['CodeListRef'])
        self.sdtm32_02['CodeListRef']= np.where(self.sdtm32_02['CodeListRef'].str.upper().isin(['MEDDRA','ISO 8601','ISO 3166-1 Alpha-3']),np.nan,self.sdtm32_02['CodeListRef'])
        self.message("Successfully filtered data")

    def read_comm(self):
        self.message("Reading COMM file")
        comm_files = [filename for filename in os.listdir(self.file_dir) if filename.endswith('_COMM.xlsx')]
        self.df_vcom01 = []
        if len(comm_files) > 0:
            file = os.path.join(self.file_dir, comm_files[0])
            self.df_vcom = pd.read_excel(file, sheet_name="COMM")
            self.df_vcom_dict = {}
            for domain in self.domains:
                self.df_com = copy.copy(self.df_vcom)
                self.df_com['Domain'] = domain
                self.df_com['Name'] = self.df_vcom['Name'].str.replace('__', domain)
                self.df_com['ProgrammerRule'] = self.df_vcom['ProgrammerRule'].str.replace('&DSN', domain)
#                 self.df_vcom01.append(self.df_com)
                self.df_vcom_dict[domain] = self.df_com
        self.message("Successfully read comm file")

    def merge_data(self):
        self.message("Merging")
        self.s_sdtm = pd.merge(self.sdtm32_02, self.df_srdm2, left_on = 'Name', right_on='V0', how = 'left')
        self.s_sdtm.head()
        self.s_sdtm.VCount = self.s_sdtm.VCount.fillna(0.0).astype(int)
        self.message("Successfully merged data")

    def f(self, row):
        val1 = ""
        val2 = ""
        if row['Name'] == row['V0']:
            if (pd.isnull(row['VCount'])):
                return val1 + val2
            if np.logical_or(row['Type'] == row['SType'], row['Label'] == row['SLabel']):
                if row['VCount'] > 1:
                    val2 = "Rename the corresponding Alias column raw variables as" + row['Name']+"|"+row['SName']
                else:
                    val1 = "Rename " + row['SName'] + " as " + row['Name']+"|"+row['SName']
            else:
                if row['VCount'] > 1:
                    val2 = "Direct move from the corresponding Alias column raw variables" +" to " + row['Name']+"|"+row['SName']
                else:
                    val1 = "Direct Move from "+row['SName']+" to " + row['Name']+"|"+row['SName']
        return val1 + ',' + val2

    def rule(self):
        self.message("Running rules engine")
        self.s_sdtm['PRuleSOrgn'] = self.s_sdtm.apply(self.f, axis=1)
        self.s_sdtm['PRule'] = np.where(self.s_sdtm['PRuleSOrgn'].isnull()
                                   ,self.s_sdtm['Name'].map(self.df_vcom.set_index('Name')['ProgrammerRule'])
                                   ,self.s_sdtm['PRuleSOrgn'].str.split('|', expand=True)[0])
        self.s_sdtm['SOrgn'] = np.where(self.s_sdtm['PRuleSOrgn'].isnull()
                                   ,self.s_sdtm['Name'].map(self.df_vcom.set_index('Name')['SRDMOrigin'])
                                   ,self.s_sdtm['PRuleSOrgn'].str.split('|', expand=True)[1])
        self.s_sdtm['SubCol'] = self.s_sdtm['Name'].map(self.df_vcom.set_index('Name')['Submission'])
        self.s_sdtm['Origin'] = self.s_sdtm['Name'].map(self.df_vcom.set_index('Name')['Origin'])
        self.s_sdtm['Length'] = self.s_sdtm['Name'].map(self.df_vcom.set_index('Name')['Length'])

        self.s_sdtm[~self.s_sdtm['SOrgn'].isnull()]
        self.message("Rules engine ran successfully")

    def apply_rule(self, row, domain, domain_df, check_col, source_col, dest_col, length, index):
        if (row["Dataset"] == domain):
            if (str(row[check_col]) == "nan"):
                filtered_df = domain_df.loc[domain_df["Name"] == row["Name"]]
                if len(filtered_df.index) > 0:
                    return filtered_df.iloc[0][dest_col]
            elif len(str(row[check_col]).split('|')) > length:
                return str(row[check_col]).split('|')[index]
        return row[source_col]

    def rule_1(self):
        self.s_sdtm['Temp'] = self.s_sdtm.apply(self.f, axis=1)
        # s_sdtm[['PRuleSOrgn', 'PRuleAlias']] = s_sdtm.Temp.str.split(',',expand=True)
        self.s_sdtm['PRuleSOrgn'] = self.s_sdtm.Temp.str.split(',',expand=True)[0]
        self.s_sdtm['PRuleSOrgn'] = self.s_sdtm['PRuleSOrgn'].replace("", np.nan)
        #s_sdtm['PRuleSOrgn'] = s_sdtm['PRuleSOrgn'].str.strip
        self.s_sdtm['PRuleAlias'] = self.s_sdtm.Temp.str.split(',',expand=True)[1]
        self.s_sdtm['PRuleAlias'] = self.s_sdtm['PRuleAlias'].replace("", np.nan)
        # s_sdtm[~s_sdtm['PRuleSOrgn'].isnull()]
        self.s_sdtm['PRule'] = ''
        self.s_sdtm['SOrgn'] = ''
        self.s_sdtm['SubCol'] = ''
        self.s_sdtm['Length'] = ''
        self.s_sdtm['SAlias'] = ''
        pd.set_option("display.max_rows", None, "display.max_columns", None)
        for domain in self.domains:
            self.domain_df = self.df_vcom_dict[domain]
#             self.s_sdtm['PRule'] = np.where(self.s_sdtm['Dataset'] == domain,
#                                        np.where(
#                                        self.s_sdtm['PRuleSOrgn'].isnull(),
#                                        self.s_sdtm['Name'].map(self.domain_df.set_index('Name')['ProgrammerRule']),
#                                        self.s_sdtm['PRuleSOrgn'].str.split('|', expand=True)[0]),
#                                        self.s_sdtm['PRule'])
            self.s_sdtm["PRule"] = self.s_sdtm.apply(self.apply_rule, axis=1, args=(domain, self.domain_df, "PRuleSOrgn", "PRule", "ProgrammerRule", 1, 0))
            self.s_sdtm["PRule"] = self.s_sdtm.apply(self.apply_rule, axis=1, args=(domain, self.domain_df, "PRuleAlias", "PRule", "ProgrammerRule", 1, 0))
            self.s_sdtm["SOrgn"] = self.s_sdtm.apply(self.apply_rule, axis=1, args=(domain, self.domain_df, "PRuleSOrgn", "SOrgn", "SRDMOrigin", 1, 1))
            self.s_sdtm['SAlias'] = self.s_sdtm.apply(self.apply_rule, axis=1, args=(domain, self.domain_df, "PRuleAlias", "SAlias", "Alias", 1, 1))
#             self.s_sdtm['SAlias'] = np.where(self.s_sdtm['Dataset'] == domain,np.where(
#                                         self.s_sdtm['PRuleAlias'].isnull(),
#                                         self.s_sdtm['Name'].map(self.domain_df.set_index('Name')['Alias']),
#                                         self.s_sdtm['PRuleAlias'].str.split('|', expand=True)[0]),
#                                         self.s_sdtm['SAlias'])
#             self.s_sdtm['SOrgn'] = np.where(self.s_sdtm['Dataset'] == domain,
#                                        np.where(
#                                        self.s_sdtm['PRuleSOrgn'].isnull(),
#                                        self.s_sdtm['Name'].map(self.domain_df.set_index('Name')['SRDMOrigin']),
#                                        self.s_sdtm['PRuleSOrgn'].str.split('|', expand=True)[1]),
#                                        self.s_sdtm['SOrgn'])
            self.s_sdtm['SubCol'] = np.where(self.s_sdtm['Dataset'] == domain,
                                        self.s_sdtm['Name'].map(self.domain_df.set_index('Name')['Submission']),
                                        self.s_sdtm['SubCol'])
            self.s_sdtm['Origin'] = np.where(self.s_sdtm['Dataset'] == domain,
                                        self.s_sdtm['Name'].map(self.domain_df.set_index('Name')['Origin']),
                                        self.s_sdtm['Origin'])
            self.s_sdtm['Length'] = np.where(self.s_sdtm['Dataset'] == domain,
                                        self.s_sdtm['Name'].map(self.domain_df.set_index('Name')['Length']),
                                        self.s_sdtm['Length'])

    def save_sheet(self):
        self.message("Exporting. Please wait...")
        book = load_workbook(self.file_path)
        self.writer = pd.ExcelWriter(self.output_path, engine="openpyxl")
        self.writer.book = book
#
        for i in range(len(self.domains)):
            domain = self.domains[i]
            self.df = pd.DataFrame(columns=['Name','Description','CodeListRef','Label','Length','Sequence',
                                                                  'Supplimentary','Comments', 'Type', 'Origin', 'Core',
                                                                  'ProgrammerRule', 'Submission', 'SRDMOrigin', 'Alias'])
            self.df1 = self.s_sdtm[self.s_sdtm['Dataset'] == domain]
            self.df['Name'] = self.sdtm32_02[self.sdtm32_02['Dataset'] == domain]['Name'].to_numpy()
#             self.df['NgCore'] = self.df['Name'].map(self.nextgen[self.nextgen['Dataset'] == domain].set_index('Name')['Core'])
            self.df['Sequence'] = self.df['Name'].map(self.sdtm32_02[self.sdtm32_02['Dataset'] == domain].set_index('Name')['Order'])
#             self.df['Description'] = self.df['Name'].map(self.df1.set_index('Name')['Description'])
            self.df['Core'] = self.df['Name'].map(self.df1.set_index('Name')['Core'])
            self.df['CodeListRef'] = self.df['Name'].map(self.df1.set_index('Name')['CodeListRef'])
            self.df['Label'] = self.df['Name'].map(self.df1.set_index('Name')['Label'])
            self.df['Type'] = self.df['Name'].map(self.df1.set_index('Name')['Type'])
            self.df.assign(Supplimentary="0")
            self.df.assign(Comments="0")
            self.df['ProgrammerRule'] = self.df['Name'].map(self.df1.set_index('Name')['PRule'])
            self.df['SRDMOrigin'] = self.df['Name'].map(self.df1.set_index('Name')['SOrgn'])
            self.df['Origin'] = self.df['Name'].map(self.df1.set_index('Name')['Origin'])
            self.df['Length'] = self.df['Name'].map(self.df1.set_index('Name')['Length'])
            self.df_copy = self.df.copy(deep=True)
            self.df_copy['NgCore'] = self.df['Name'].map(self.sdtm32_02[self.sdtm32_02['Dataset'] == domain].set_index('Name')['Core'])

            # Save with formatting
            self.df.to_excel(self.writer, sheet_name=domain, index=False)
            worksheet = self.writer.sheets[domain]
            self.format_all_cells(worksheet, domain, "A", 20)
            self.format_all_cells(worksheet, domain, "B", 40)
            self.format_all_cells(worksheet, domain, "C", 20)
            self.format_all_cells(worksheet, domain, "D", 20)
            self.format_all_cells(worksheet, domain, "E", 12)
            self.format_all_cells(worksheet, domain, "F", 12)
            self.format_all_cells(worksheet, domain, "G", 15)
            self.format_all_cells(worksheet, domain, "H", 12)
            self.format_all_cells(worksheet, domain, "I", 12)
            self.format_all_cells(worksheet, domain, "J", 12)
            self.format_all_cells(worksheet, domain, "K", 12)
            self.format_all_cells(worksheet, domain, "L", 40)
            self.format_all_cells(worksheet, domain, "M", 12)
            self.format_all_cells(worksheet, domain, "N", 20)
            self.format_all_cells(worksheet, domain, "O", 12)
#                 self.format_all_cells(worksheet, domain, "P", 20)
            self.format_header_cells(worksheet)
            self.format_core(worksheet)

        self.writer.save()
        self.writer.close()
        self.message("Successfully exported sheet")

    def format_all_cells(self, sheet, domain, col_index, width):
        cell_style = NamedStyle(name=domain+col_index, alignment=Alignment(wrap_text=True))
        for cell in sheet[col_index]:
            cell.style = cell_style
        sheet.column_dimensions[col_index].width = width

    def format_header_cells(self, sheet):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
        for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
              cell.fill = PatternFill(start_color="eeeeee", end_color="dddddd", fill_type = "solid")
              cell.font = Font(bold=True)
              cell.border = thin_border

    def format_core(self, sheet):
        for i, row in self.df_copy.iterrows():
            if row["NgCore"] == "Model Permissible":
                cell = sheet.cell(row=i, column=11)
                cell.fill = PatternFill(start_color="eeb3a5", end_color="eeb3a5", fill_type = "solid")

    def process(self):
        try:
            self.get_domains()
    #         self.read_nextgen()
    #         self.read_sdtm_ig()
            self.read_sdtm_meta()
            self.get_unique_domains_from_sdtm()
    #         self.append_data_for_all_domains()
    #         self.set_desired_columns()
            self.read_comm()
            self.merge_data()
    #             self.rule()
            self.rule_1()
            self.save_sheet()
        except Exception as e:
            self.message("[ERROR] "+str(e))
